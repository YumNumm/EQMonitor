import pyexcel as pe
import openpyxl
import argparse

def convert_xls_to_xlsx(input_file, output_file):
  pe.save_book_as(file_name=input_file, dest_file_name=input_file + ".xlsx")
  # ついでに CSV にも変換しておく
  wb = openpyxl.load_workbook(input_file + ".xlsx")
  for sheet in wb.sheetnames:
    ws = wb[sheet]
    with open(output_file + "_" + sheet + ".csv", "w") as f:
      for row in ws.iter_rows():
        f.write(",".join([str(cell.value) for cell in row]) + "\n")
  



if __name__ == "__main__":
  # コマンドライン引数を受け取る
  parser = argparse.ArgumentParser()
  # -i
  parser.add_argument("-i", "--input_file", help="input file path", required=True)
  # -o
  parser.add_argument("-o", "--output_file", help="output file path", required=True)
  args = parser.parse_args()

  # xlsファイルをxlsxファイルに変換する
  convert_xls_to_xlsx(args.input_file, args.output_file)